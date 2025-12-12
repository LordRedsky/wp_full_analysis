import io
from typing import Tuple, Dict, Any

from exact_match import (
    _read_excel_both,
    _normalize_str,
    _find_header_col,
    _is_unfilled,
    _append_ket,
    _strip_titles_and_prefixes,
    _contains_title,
    _contains_prefix,
    _to_int,
    _parse_currency_to_int,
    _set_number,
    _set_padded_number,
    RED_FILL,
    BLUE_FILL,
    GREEN_FILL,
    GREY_FILL,
)


def _is_same_fill(cell, fill) -> bool:
    f = getattr(cell, "fill", None)
    if f is None or getattr(f, "patternType", None) is None:
        return False
    try:
        return f.patternType == "solid" and f.start_color.rgb == fill.start_color.rgb
    except Exception:
        return False


def _set_ket_numbered(ws, row: int, col: int, messages):
    cleaned = [str(m).strip() for m in messages if m is not None and str(m).strip() != ""]
    if not cleaned:
        ws.cell(row=row, column=col, value="")
        return
    lines = [f"{i+1}. {msg}" for i, msg in enumerate(cleaned)]
    ws.cell(row=row, column=col, value="\n".join(lines))


def _set_padded_text(ws, row: int, col: int, value: Any, width: int):
    n = _to_int(value)
    if n is None:
        s = str(value).strip()
        digits = "".join(ch for ch in s if ch.isdigit())
        if digits:
            n = int(digits)
    if n is None:
        ws.cell(row=row, column=col, value="")
    else:
        s = str(n).zfill(width)
        ws.cell(row=row, column=col, value=s)
        ws.cell(row=row, column=col).number_format = "@"
#ff
def process_deep_analysis(
    wajib_pajak_file: bytes,
    sertifikat_file: bytes,
    kendali_file: bytes,
) -> Tuple[bytes, bytes, bytes, Dict[str, int]]:
    wp_df, wp_wb = _read_excel_both(wajib_pajak_file)
    cert_df, cert_wb = _read_excel_both(sertifikat_file)
    kendali_df, _ = _read_excel_both(kendali_file)

    wp_ws = wp_wb.active
    cert_ws = cert_wb.active

    for col in ["NIB", "Luas"]:
        if col in cert_df.columns:
            cert_df[col] = cert_df[col].apply(_to_int)

    for col in ["NIB", "LUASTERTUL", "RPBULAT"]:
        if col in kendali_df.columns:
            if col == "NIB":
                kendali_df[col] = kendali_df[col].apply(_to_int)
            elif col == "LUASTERTUL":
                kendali_df[col] = kendali_df[col].apply(_to_int)
            elif col == "RPBULAT":
                kendali_df[col] = kendali_df[col].apply(_parse_currency_to_int)

    wp_name_col = _find_header_col(wp_ws, "NAMA_WP")
    wp_jlh_col = _find_header_col(wp_ws, "JENIS_LHN")
    wp_kls_col = _find_header_col(wp_ws, "KELAS_BUMI")
    wp_nib1_col = _find_header_col(wp_ws, "NIB_1")
    wp_kdznt_col = _find_header_col(wp_ws, "KD_ZNT")
    wp_njop_col = _find_header_col(wp_ws, "NJOP_BUMI")
    wp_luas_col = _find_header_col(wp_ws, "LUAS_BUMI")

    cert_nama_col = _find_header_col(cert_ws, "Nama")
    cert_nib_col = _find_header_col(cert_ws, "NIB")
    cert_luas_col = _find_header_col(cert_ws, "Luas")

    try:
        cert_ket_col = _find_header_col(cert_ws, "KETERANGAN")
    except KeyError:
        cert_ket_col = cert_ws.max_column + 1
        cert_ws.cell(row=1, column=cert_ket_col, value="KETERANGAN")

    try:
        wp_ket_col = _find_header_col(wp_ws, "KETERANGAN")
    except KeyError:
        wp_ket_col = wp_ws.max_column + 1
        wp_ws.cell(row=1, column=wp_ket_col, value="KETERANGAN")

    base_to_cert_rows: Dict[str, list] = {}
    for r in range(2, cert_ws.max_row + 1):
        nama_cell = cert_ws.cell(row=r, column=cert_nama_col)
        if not _is_unfilled(nama_cell):
            continue
        nama_val = _normalize_str(nama_cell.value)
        if not nama_val:
            continue
        nib_val = cert_ws.cell(row=r, column=cert_nib_col).value
        luas_val = cert_ws.cell(row=r, column=cert_luas_col).value
        item = {"row": r, "NIB": _to_int(nib_val), "Luas": _to_int(luas_val), "Nama": nama_val}
        base_key = _strip_titles_and_prefixes(nama_val)
        if base_key:
            base_to_cert_rows.setdefault(base_key, []).append(item)

    kendali_by_nib: Dict[int, Dict[str, Any]] = {}
    if "NIB" in kendali_df.columns:
        for _, row in kendali_df.iterrows():
            nib = _to_int(row.get("NIB"))
            if nib is None:
                continue
            kendali_by_nib[nib] = {
                "RPBULAT": _parse_currency_to_int(row.get("RPBULAT")),
                "JENIS_ZONA": _normalize_str(row.get("JENIS_ZONA")),
                "KELAS_BUMI": _normalize_str(row.get("KELAS_BUMI")),
                "KD_ZNT": _normalize_str(row.get("KD_ZNT")),
                "LUASTERTUL": _to_int(row.get("LUASTERTUL")),
            }

    used_nibs = set()
    for r in range(2, wp_ws.max_row + 1):
        existing_nib = _to_int(wp_ws.cell(row=r, column=wp_nib1_col).value)
        if existing_nib is not None:
            used_nibs.add(existing_nib)

    processed_grey = 0
    updated_count = 0
    nib_not_found_count = 0
    name_not_found_count = 0
    duplicate_nib_count = 0

    for r in range(2, wp_ws.max_row + 1):
        name_cell = wp_ws.cell(row=r, column=wp_name_col)
        if not _is_same_fill(name_cell, GREY_FILL):
            continue
        nama_wp_raw = name_cell.value
        nama_wp = _normalize_str(nama_wp_raw)
        if not nama_wp:
            continue
        processed_grey += 1
        has_decor = _contains_title(nama_wp) or _contains_prefix(nama_wp)
        base_wp = _strip_titles_and_prefixes(nama_wp) if has_decor else _strip_titles_and_prefixes(nama_wp)
        candidates = base_to_cert_rows.get(base_wp)
        if not candidates:
            wp_ws.cell(row=r, column=wp_name_col).fill = RED_FILL
            wp_ws.cell(row=r, column=wp_ket_col, value="Nama Wajib Pajak tidak ditemukan di sertifikat")
            name_not_found_count += 1
            continue
        chosen_item = None
        chosen_kendali = None
        for it in candidates:
            row_chk = it["row"]
            if not _is_unfilled(cert_ws.cell(row=row_chk, column=cert_nama_col)):
                continue
            nib = it["NIB"]
            if nib is None:
                continue
            if nib in used_nibs:
                continue
            kendali = kendali_by_nib.get(nib)
            if kendali:
                chosen_item = it
                chosen_kendali = kendali
                break
        if chosen_item is None:
            any_white = [it for it in candidates if _is_unfilled(cert_ws.cell(row=it["row"], column=cert_nama_col))]
            if not any_white:
                wp_ws.cell(row=r, column=wp_name_col).fill = RED_FILL
                wp_ws.cell(row=r, column=wp_ket_col, value="Nama Wajib Pajak tidak ditemukan di sertifikat")
                name_not_found_count += 1
                continue
            all_used = [it for it in any_white if it["NIB"] in used_nibs]
            if all_used and len(all_used) == len(any_white):
                wp_ws.cell(row=r, column=wp_name_col).fill = RED_FILL
                wp_ws.cell(row=r, column=wp_ket_col, value="NIB telah digunakan pada baris lain")
                duplicate_nib_count += 1
                continue
            it = any_white[0]
            cert_ws.cell(row=it["row"], column=cert_nama_col).fill = RED_FILL
            wp_ws.cell(row=r, column=wp_name_col).fill = RED_FILL
            _append_ket(cert_ws, it["row"], cert_ket_col, "NIB tidak ditemukan pada form kendali")
            wp_ws.cell(row=r, column=wp_ket_col, value="NIB tidak ditemukan pada form kendali")
            nib_not_found_count += 1
            continue

        it = chosen_item
        kendali = chosen_kendali

        wp_ws.cell(row=r, column=wp_jlh_col, value=kendali["JENIS_ZONA"])
        _set_padded_text(wp_ws, r, wp_kls_col, kendali["KELAS_BUMI"], 3)
        _set_padded_text(wp_ws, r, wp_nib1_col, it["NIB"], 5)
        wp_ws.cell(row=r, column=wp_kls_col).number_format = "000"
        wp_ws.cell(row=r, column=wp_nib1_col).number_format = "00000"
        wp_ws.cell(row=r, column=wp_kdznt_col, value=kendali["KD_ZNT"])
        njop = kendali["RPBULAT"]
        if njop is not None:
            _set_number(wp_ws, r, wp_njop_col, njop)
        else:
            wp_ws.cell(row=r, column=wp_njop_col, value=njop)
        messages = []

        luas_tertul = kendali["LUASTERTUL"]
        old_wp_luas = wp_ws.cell(row=r, column=wp_luas_col).value
        old_norm = _to_int(old_wp_luas)
        tertul_norm = _to_int(luas_tertul)
        if tertul_norm is not None and old_norm != tertul_norm:
            _set_number(wp_ws, r, wp_luas_col, tertul_norm)
            wp_ws.cell(row=r, column=wp_luas_col).fill = GREEN_FILL
            messages.append(f"Telah dilakukan penyesuaian luas bumi sesuai data sertipikat. {old_wp_luas} -> {luas_tertul}")

        old_name = wp_ws.cell(row=r, column=wp_name_col).value
        new_name = it["Nama"]
        wp_ws.cell(row=r, column=wp_name_col, value=new_name)
        wp_ws.cell(row=r, column=wp_name_col).fill = GREEN_FILL
        messages.append(f"Telah dilakukan penyesuaian nama sesuai sertifikat. {old_name} -> {new_name}")
        _set_ket_numbered(wp_ws, r, wp_ket_col, messages)

        for c in [cert_nama_col, cert_nib_col, cert_luas_col]:
            cert_ws.cell(row=it["row"], column=c).fill = BLUE_FILL
        _append_ket(cert_ws, it["row"], cert_ket_col, "Data cocok")
        used_nibs.add(it["NIB"])
        updated_count += 1

    wp_out = io.BytesIO()
    team_out = io.BytesIO()
    cert_out = io.BytesIO()
    wp_wb.save(wp_out)
    from openpyxl import Workbook
    team_wb = Workbook()
    team_ws = team_wb.active
    for c in range(1, wp_ws.max_column + 1):
        team_ws.cell(row=1, column=c, value=wp_ws.cell(row=1, column=c).value)
    tr = 2
    for r in range(2, wp_ws.max_row + 1):
        name_cell = wp_ws.cell(row=r, column=wp_name_col)
        if _is_same_fill(name_cell, RED_FILL):
            continue
        for c in range(1, wp_ws.max_column + 1):
            src = wp_ws.cell(row=r, column=c)
            dst = team_ws.cell(row=tr, column=c, value=src.value)
            try:
                dst.number_format = src.number_format
            except Exception:
                pass
            try:
                dst.fill = src.fill
            except Exception:
                pass
            if c == wp_name_col or c == wp_luas_col:
                if _is_same_fill(src, GREEN_FILL):
                    dst.fill = GREEN_FILL
                elif _is_same_fill(src, GREY_FILL):
                    dst.fill = GREY_FILL
        tr += 1
    team_wb.save(team_out)
    cert_wb.save(cert_out)
    stats = {
        "processed_grey": processed_grey,
        "updated": updated_count,
        "matched": updated_count,
        "nib_not_found": nib_not_found_count,
        "name_not_found": name_not_found_count,
        "duplicate_nib": duplicate_nib_count,
    }
    return wp_out.getvalue(), team_out.getvalue(), cert_out.getvalue(), stats
