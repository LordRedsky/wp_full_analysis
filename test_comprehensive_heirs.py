#!/usr/bin/env python3
"""
Comprehensive test for heirs detection functionality
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from analysis_sert import process_certificate_analysis


def create_comprehensive_test_sertifikat():
    """Create a test sertifikat Excel file with specific cases for heirs detection"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add headers
    ws.cell(row=1, column=1, value="Nama")
    ws.cell(row=1, column=2, value="NIB")
    ws.cell(row=1, column=3, value="Luas")

    # Row 2: Single name with white-filled NIB and NAMA - should NOT be yellow
    ws.cell(row=2, column=1, value="John Doe")
    ws.cell(row=2, column=2, value="12345")
    ws.cell(row=2, column=3, value=100)

    # Row 3: Multiple names indicating heirs, both NIB and NAMA white-filled - should be yellow
    ws.cell(row=3, column=1, value="HAFIDIN, A.MA.PD.I, Drs. Amin, Wa Uceng, La Deli")
    ws.cell(row=3, column=2, value="67890")
    ws.cell(row=3, column=3, value=200)

    # Row 4: Empty NIB (should be marked as red)
    ws.cell(row=4, column=1, value="Jane Smith")
    ws.cell(row=4, column=2, value="")  # Empty NIB
    ws.cell(row=4, column=3, value=150)

    # Row 5: Multiple names with white-filled NIB and NAMA - should be yellow
    ws.cell(row=5, column=1, value="Siti, H. Ahmad, Drs. Budi")
    ws.cell(row=5, column=2, value="11111")
    ws.cell(row=5, column=3, value=300)

    # Row 6: Single name with title, both white-filled - should NOT be yellow
    ws.cell(row=6, column=1, value="Drs. Joko")
    ws.cell(row=6, column=2, value="22222")
    ws.cell(row=6, column=3, value=250)

    # Row 7: Multiple names with colored NIB - should NOT be processed for heirs
    # Let's set a background color for this NIB cell to test the condition
    ws.cell(row=7, column=1, value="Ali, Budi, Charlie")
    ws.cell(row=7, column=2, value="33333")
    ws.cell(row=7, column=3, value=400)
    # Add a colored fill to the NIB cell (not white)
    ws.cell(row=7, column=2).fill = PatternFill(start_color="FFC0C0C0", end_color="FFC0C0C0", fill_type="solid")

    # Row 8: Multiple names with colored NAMA - should NOT be processed for heirs
    ws.cell(row=8, column=1, value="Dedi, Eka, Fani")
    ws.cell(row=8, column=2, value="44444")
    ws.cell(row=8, column=3, value=500)
    # Add a colored fill to the NAMA cell (not white)
    ws.cell(row=8, column=1).fill = PatternFill(start_color="FFC0C0C0", end_color="FFC0C0C0", fill_type="solid")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def test_comprehensive_heirs_detection():
    """Test the certificate analysis with comprehensive heirs detection"""
    print("Creating comprehensive test sertifikat file...")
    test_file = create_comprehensive_test_sertifikat()

    print("Running certificate analysis with heirs detection...")
    result = process_certificate_analysis(test_file)

    print("Certificate analysis completed successfully!")
    print(f"Data diproses: {result['data_diproses']}")
    print(f"NIB duplikat: {result['nib_duplikat']}")
    print(f"Jumlah Ahli Waris: {result['jumlah_ahli_waris']}")

    # Load the result to check colors
    from openpyxl import load_workbook
    result_wb = load_workbook(io.BytesIO(result['workbook_bytes']), data_only=True)
    result_ws = result_wb.active

    print("\nAnalyzing results in detail:")
    for r in range(2, result_ws.max_row + 1):
        nama = result_ws.cell(row=r, column=1).value
        nib = result_ws.cell(row=r, column=2).value
        luas = result_ws.cell(row=r, column=3).value

        # Check the fill color
        nama_cell = result_ws.cell(row=r, column=1)
        nib_cell = result_ws.cell(row=r, column=2)
        luas_cell = result_ws.cell(row=r, column=3)

        # Determine if cells have white/empty fill (unfilled)
        from analysis_sert import _is_unfilled
        nama_is_unfilled = _is_unfilled(nama_cell)
        nib_is_unfilled = _is_unfilled(nib_cell)

        # Determine actual colors
        is_red = (hasattr(nama_cell.fill, 'start_color') and 
                 nama_cell.fill.start_color.rgb == "FFFF0000") if hasattr(nama_cell.fill, 'start_color') else False
        is_yellow = (hasattr(nama_cell.fill, 'start_color') and 
                    nama_cell.fill.start_color.rgb == "FFFFFF00") if hasattr(nama_cell.fill, 'start_color') else False

        print(f"Row {r}: Nama='{nama}', NIB='{nib}', Luas={luas}")
        print(f"  NIB unfilled: {nib_is_unfilled}, NAMA unfilled: {nama_is_unfilled}")
        print(f"  Is Red: {is_red}, Is Yellow: {is_yellow}")

    print(f"\nTotal heirs detected: {result['jumlah_ahli_waris']}")
    
    # Verify expected results
    print("\nExpected results:")
    print("- Row 2: Single name, both white-filled -> Should NOT be yellow")
    print("- Row 3: Multiple names (heirs), both white-filled -> Should be yellow")
    print("- Row 4: Empty NIB -> Should be red with 'NIB kosong'")
    print("- Row 5: Multiple names (heirs), both white-filled -> Should be yellow")
    print("- Row 6: Single name with title, both white-filled -> Should NOT be yellow")
    print("- Row 7: Multiple names, but NIB has color -> Should NOT be yellow (not processed for heirs)")
    print("- Row 8: Multiple names, but NAMA has color -> Should NOT be yellow (not processed for heirs)")
    
    # Count yellow rows (potential heirs)
    yellow_count = 0
    for r in range(2, result_ws.max_row + 1):
        nama_cell = result_ws.cell(row=r, column=1)
        is_yellow = (hasattr(nama_cell.fill, 'start_color') and 
                    nama_cell.fill.start_color.rgb == "FFFFFF00") if hasattr(nama_cell.fill, 'start_color') else False
        if is_yellow:
            yellow_count += 1
    
    print(f"\nActual yellow rows (heirs detected): {yellow_count}")
    print(f"Reported heirs count: {result['jumlah_ahli_waris']}")
    
    # Expected: Rows 3 and 5 should be yellow (2 heirs)
    expected_heirs = 2
    assert result['jumlah_ahli_waris'] == expected_heirs, f"Expected {expected_heirs} heirs, but found {result['jumlah_ahli_waris']}"
    
    print(f"\nTest passed! Found {result['jumlah_ahli_waris']} records with heirs as expected.")


if __name__ == "__main__":
    test_comprehensive_heirs_detection()
    print("\nComprehensive test passed successfully!")