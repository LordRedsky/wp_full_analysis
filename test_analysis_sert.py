import io
import pandas as pd
from openpyxl import Workbook
from analysis_sert import process_certificate_analysis

def create_test_sertifikat():
    """Create a test sertifikat Excel file with white-filled NIB values for testing"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add headers
    ws.cell(row=1, column=1, value="Nama")
    ws.cell(row=1, column=2, value="NIB")
    ws.cell(row=1, column=3, value="Luas")

    # Add some test data with white fills
    # Row 2: Valid NIB (first occurrence should NOT be marked red)
    ws.cell(row=2, column=1, value="John Doe")
    ws.cell(row=2, column=2, value="12345")
    ws.cell(row=2, column=3, value=100)

    # Row 3: Empty NIB (should be marked as red)
    ws.cell(row=3, column=1, value="Jane Smith")
    ws.cell(row=3, column=2, value="")  # Empty NIB
    ws.cell(row=3, column=3, value=200)

    # Row 4: Duplicate NIB (should be marked as red - according to updated requirement)
    ws.cell(row=4, column=1, value="Bob Johnson")
    ws.cell(row=4, column=2, value="12345")  # Duplicate of row 2 - this should be marked red
    ws.cell(row=4, column=3, value=150)

    # Row 5: Valid NIB
    ws.cell(row=5, column=1, value="Alice Brown")
    ws.cell(row=5, column=2, value="67890")
    ws.cell(row=5, column=3, value=300)

    # Row 6: Another duplicate (should be marked as red)
    ws.cell(row=6, column=1, value="Charlie Wilson")
    ws.cell(row=6, column=2, value="67890")  # Duplicate of row 5 - this should be marked red
    ws.cell(row=6, column=3, value=250)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def test_certificate_analysis():
    print("Creating test sertifikat file...")
    test_file = create_test_sertifikat()

    print("Running certificate analysis...")
    result = process_certificate_analysis(test_file)

    print("Certificate analysis completed successfully!")

    print(f"Data diproses: {result['data_diproses']}")
    print(f"NIB duplikat: {result['nib_duplikat']}")

    # Let's verify by loading the result and checking some values
    result_wb = load_workbook(io.BytesIO(result['workbook_bytes']), data_only=True)
    result_ws = result_wb.active

    print("\nAnalyzing results:")
    print("(Based on updated requirement: only subsequent duplicates should be marked red, not first occurrence)")
    for r in range(2, min(7, result_ws.max_row + 1)):  # Check up to row 6
        nama = result_ws.cell(row=r, column=1).value
        nib = result_ws.cell(row=r, column=2).value
        luas = result_ws.cell(row=r, column=3).value

        # Check the fill color
        nama_fill = result_ws.cell(row=r, column=1).fill
        nib_fill = result_ws.cell(row=r, column=2).fill
        luas_fill = result_ws.cell(row=r, column=3).fill

        is_red = (nama_fill.start_color.rgb == "FFFF0000" if hasattr(nama_fill, 'start_color') else False)

        print(f"Row {r}: Nama={nama}, NIB={nib}, Luas={luas}, Is Red: {is_red}")

    # Test the expected results:
    print("\nExpected results based on updated requirements:")
    print("- Row 2: NIB=12345 (first occurrence) -> Should NOT be red")
    print("- Row 3: NIB=None (empty) -> Should be red with 'NIB kosong'")
    print("- Row 4: NIB=12345 (duplicate) -> Should be red with 'NIB memiliki duplikate'")
    print("- Row 5: NIB=67890 (first occurrence) -> Should NOT be red")
    print("- Row 6: NIB=67890 (duplicate) -> Should be red with 'NIB memiliki duplikate'")

    print("\nTest completed!")

if __name__ == "__main__":
    from openpyxl import load_workbook
    test_certificate_analysis()