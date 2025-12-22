import pandas as pd

try:
    df = pd.read_excel('Sertifikat_analysed.xlsx')
    print('File loaded successfully')
    print('Column names:', df.columns.tolist())
    print('Total rows:', len(df))
    print()

    # Check row 19 (index 18) - the last row
    if len(df) >= 19:
        print('Row 19 (index 18):')
        for col in df.columns:
            val = df.iloc[18][col]
            print(f'  {col}: {val}')
        print()

        # Check if there might be multiple names in the Nama column
        nama_val = df.iloc[18]['Nama']
        print(f'Nama in row 19: \"{nama_val}\"')

        # Test our parsing function on this value
        from analysis_sert import _parse_names
        parsed = _parse_names(nama_val)
        print(f'Parsed names: {parsed}')
        print(f'Number of parsed names: {len(parsed)}')

    if len(df) >= 20:
        print('Row 20 (index 19):')
        for col in df.columns:
            print(f'  {col}: {df.iloc[19][col]}')
    else:
        print(f'Dataframe only has {len(df)} rows')

except Exception as e:
    print(f'Error reading file: {e}')

    # Try with openpyxl to see if we can access it
    try:
        from openpyxl import load_workbook
        wb = load_workbook('Sertifikat_analysed.xlsx')
        ws = wb.active
        print(f'Worksheet has {ws.max_row} rows and {ws.max_column} columns')

        # Print headers
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(row=1, column=col).value)
        print('Headers:', headers)

        # Print row 19 (last row)
        if ws.max_row >= 19:
            print('\nRow 19:')
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=19, column=col)
                print(f'  {headers[col-1]}: {cell.value}')

            if ws.max_row >= 20:
                print('\nRow 20:')
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=20, column=col)
                    print(f'  {headers[col-1]}: {cell.value}')
    except Exception as e2:
        print(f'Error with openpyxl: {e2}')