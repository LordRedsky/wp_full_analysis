import io
from typing import Tuple, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
BLUE_FILL = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
GREY_FILL = PatternFill(start_color="FFBFBFBF", end_color="FFBFBFBF", fill_type="solid")


def _read_excel_both(file_bytes: bytes) -> Tuple[pd.DataFrame, Any]:
    df = pd.read_excel(io.BytesIO(file_bytes))
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    return df, wb


def _is_unfilled(cell) -> bool:
    return cell.fill is None or cell.fill.patternType is None


def _to_int(x: Any):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    try:
        if isinstance(x, (int, float)):
            return int(float(x))
        s = str(x).strip()
        s = s.replace("Rp", "").replace(" ", "").replace(",", "")
        if s.count(".") <= 1:
            try:
                return int(float(s))
            except Exception:
                pass
        digits = "".join(ch for ch in s if ch.isdigit())
        return int(digits) if digits else None
    except Exception:
        return None


def _find_header_col(ws, header_name: str) -> int:
    header_name_norm = header_name.lower().strip()
    for cell in ws[1]:
        if str(cell.value).lower().strip() == header_name_norm:
            return cell.column
    raise KeyError(f"Kolom '{header_name}' tidak ditemukan")


def _append_ket(ws, row: int, col: int, msg: str) -> None:
    existing = ws.cell(row=row, column=col).value
    if existing is None or str(existing).strip() == "":
        ws.cell(row=row, column=col, value=f"1. {msg}")
    else:
        lines = [l for l in str(existing).splitlines() if l.strip()]
        n = len(lines) + 1
        ws.cell(row=row, column=col, value=str(existing) + "\n" + f"{n}. {msg}")


def process_certificate_analysis(
    sertifikat_file: bytes,
) -> bytes:
    """
    Performs certificate analysis on the given sertifikat file.

    Requirements:
    - Only analyzes white/empty-filled NIB columns
    - Checks if NIB values exist (not empty)
    - Checks for duplicate NIB values
    - Colors rows red if NIB is empty or duplicate
    - Only subsequent duplicates (not first occurrence) are marked red
    - Adds appropriate descriptions
    """
    sertifikat_df, sertifikat_wb = _read_excel_both(sertifikat_file)
    sertifikat_ws = sertifikat_wb.active

    # Find required columns
    nama_col = _find_header_col(sertifikat_ws, "Nama")
    nib_col = _find_header_col(sertifikat_ws, "NIB")
    luas_col = _find_header_col(sertifikat_ws, "Luas")

    # Find or create KETERANGAN column
    try:
        ket_col = _find_header_col(sertifikat_ws, "KETERANGAN")
    except KeyError:
        ket_col = sertifikat_ws.max_column + 1
        sertifikat_ws.cell(row=1, column=ket_col, value="KETERANGAN")

    # Collect white-filled NIB values and their rows in order
    white_nib_data = []

    for r in range(2, sertifikat_ws.max_row + 1):
        nib_cell = sertifikat_ws.cell(row=r, column=nib_col)

        # Only process white/empty-filled NIB cells
        if _is_unfilled(nib_cell):
            nib_value = _to_int(nib_cell.value)
            white_nib_data.append({
                'row': r,
                'nib_value': nib_value,
                'cell_obj': nib_cell
            })

    # Track which NIB values have been seen (to identify subsequent duplicates)
    seen_nibs = set()
    duplicate_rows = set()  # Track rows that should be marked as red due to duplication

    for data in white_nib_data:
        nib_value = data['nib_value']

        # Check if NIB is empty/None
        if nib_value is None or pd.isna(nib_value):
            # Mark row as red and add description
            row = data['row']
            sertifikat_ws.cell(row=row, column=nama_col).fill = RED_FILL
            sertifikat_ws.cell(row=row, column=nib_col).fill = RED_FILL
            sertifikat_ws.cell(row=row, column=luas_col).fill = RED_FILL
            _append_ket(sertifikat_ws, row, ket_col, "NIB kosong")
        else:
            # Check if this is a duplicate (not the first occurrence)
            if nib_value in seen_nibs:
                # This is a duplicate occurrence, mark it as red
                row = data['row']
                duplicate_rows.add(row)
                sertifikat_ws.cell(row=row, column=nama_col).fill = RED_FILL
                sertifikat_ws.cell(row=row, column=nib_col).fill = RED_FILL
                sertifikat_ws.cell(row=row, column=luas_col).fill = RED_FILL
                _append_ket(sertifikat_ws, row, ket_col, "NIB memiliki duplikate")
            else:
                # First occurrence, just add to seen set
                seen_nibs.add(nib_value)

    # Process all white-filled NIB cells for empty check and duplicate marking
    for data in white_nib_data:
        row = data['row']
        nib_value = data['nib_value']

        # Empty check is already handled above
        # Duplicate marking is already handled above

    # Save the result
    sertifikat_out = io.BytesIO()
    sertifikat_wb.save(sertifikat_out)

    return sertifikat_out.getvalue()


if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python analysis_sert.py <input_sertifikat_file> <output_sertifikat_file>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    # Read the input file
    with open(input_file, 'rb') as f:
        sertifikat_bytes = f.read()

    # Process the certificate analysis
    result_bytes = process_certificate_analysis(sertifikat_bytes)

    # Write the result to output file
    with open(output_file, 'wb') as f:
        f.write(result_bytes)

    print(f"Certificate analysis completed. Output saved to {output_file}")