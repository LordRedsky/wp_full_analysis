#!/usr/bin/env python3
"""
Test script to verify the heirs detection functionality in analysis_sert.py
"""

import io
import pandas as pd
from openpyxl import Workbook
from analysis_sert import process_certificate_analysis, _parse_names


def create_test_sertifikat_with_heirs():
    """Create a test sertifikat Excel file with examples of heirs"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add headers
    ws.cell(row=1, column=1, value="Nama")
    ws.cell(row=1, column=2, value="NIB")
    ws.cell(row=1, column=3, value="Luas")

    # Row 2: Single name (no heirs) - should NOT be marked yellow
    ws.cell(row=2, column=1, value="John Doe")
    ws.cell(row=2, column=2, value="12345")
    ws.cell(row=2, column=3, value=100)

    # Row 3: Multiple names indicating heirs - should be marked yellow
    ws.cell(row=3, column=1, value="HAFIDIN, A.MA.PD.I, Drs. Amin, Wa Uceng, La Deli")
    ws.cell(row=3, column=2, value="67890")
    ws.cell(row=3, column=3, value=200)

    # Row 4: Empty NIB (should be marked as red)
    ws.cell(row=4, column=1, value="Jane Smith")
    ws.cell(row=4, column=2, value="")  # Empty NIB
    ws.cell(row=4, column=3, value=150)

    # Row 5: Another example of multiple names - should be marked yellow
    ws.cell(row=5, column=1, value="Siti, H. Ahmad, Drs. Budi")
    ws.cell(row=5, column=2, value="11111")
    ws.cell(row=5, column=3, value=300)

    # Row 6: Single name with title - should NOT be marked yellow
    ws.cell(row=6, column=1, value="Drs. Joko")
    ws.cell(row=6, column=2, value="22222")
    ws.cell(row=6, column=3, value=250)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def test_name_parsing():
    """Test the name parsing function"""
    print("Testing name parsing function...")
    
    test_cases = [
        ("HAFIDIN, A.MA.PD.I, Drs. Amin, Wa Uceng, La Deli", 4),  # Should extract 4 names: HAFIDIN, Amin, Uceng, Deli
        ("Siti, H. Ahmad, Drs. Budi", 3),  # Should extract 3 names: Siti, Ahmad, Budi
        ("Drs. Joko", 1),  # Should extract 1 name: Joko
        ("John Doe", 1),  # Should extract 1 name: John
        ("", 0),  # Should extract 0 names
        (None, 0),  # Should extract 0 names
    ]
    
    for name_input, expected_count in test_cases:
        result = _parse_names(name_input)
        print(f"Input: '{name_input}' -> Parsed: {result} (Count: {len(result)}, Expected: {expected_count})")
        assert len(result) == expected_count, f"Failed for input '{name_input}'. Expected {expected_count}, got {len(result)}"
    
    print("All name parsing tests passed!\n")


def test_certificate_analysis_with_heirs():
    """Test the certificate analysis with heirs detection"""
    print("Creating test sertifikat file with heirs...")
    test_file = create_test_sertifikat_with_heirs()

    print("Running certificate analysis with heirs detection...")
    result = process_certificate_analysis(test_file)

    print("Certificate analysis completed successfully!")
    print(f"Data diproses: {result['data_diproses']}")
    print(f"NIB duplikat: {result['nib_duplikat']}")
    print(f"Jumlah Ahli Waris: {result['jumlah_ahli_waris']}")

    # Load the result to check colors
    from openpyxl import load_workbook
    result_wb = load_workbook(io.BytesIO(result['workbook_bytes']), data_only=True)
    result_ws = result_wb.active

    print("\nAnalyzing results:")
    for r in range(2, min(7, result_ws.max_row + 1)):  # Check up to row 6
        nama = result_ws.cell(row=r, column=1).value
        nib = result_ws.cell(row=r, column=2).value
        luas = result_ws.cell(row=r, column=3).value

        # Check the fill color
        nama_fill = result_ws.cell(row=r, column=1).fill
        nib_fill = result_ws.cell(row=r, column=2).fill
        luas_fill = result_ws.cell(row=r, column=3).fill

        # Determine color based on fill pattern
        is_red = (hasattr(nama_fill, 'start_color') and nama_fill.start_color.rgb == "FFFF0000") if hasattr(nama_fill, 'start_color') else False
        is_yellow = (hasattr(nama_fill, 'start_color') and nama_fill.start_color.rgb == "FFFFFF00") if hasattr(nama_fill, 'start_color') else False

        print(f"Row {r}: Nama='{nama}', NIB='{nib}', Luas={luas}, Is Red: {is_red}, Is Yellow: {is_yellow}")

    # Verify expected results
    print("\nExpected results:")
    print("- Row 2: Single name -> Should NOT be yellow")
    print("- Row 3: Multiple names (heirs) -> Should be yellow with 'Data Ahli Waris'")
    print("- Row 4: Empty NIB -> Should be red with 'NIB kosong'")
    print("- Row 5: Multiple names (heirs) -> Should be yellow with 'Data Ahli Waris'")
    print("- Row 6: Single name with title -> Should NOT be yellow")
    
    # Check if the number of heirs detected matches expectation
    expected_heirs = 2  # Rows 3 and 5 should be marked as heirs
    assert result['jumlah_ahli_waris'] == expected_heirs, f"Expected {expected_heirs} heirs, but found {result['jumlah_ahli_waris']}"
    
    print(f"\nTest passed! Found {result['jumlah_ahli_waris']} records with heirs as expected.")


if __name__ == "__main__":
    test_name_parsing()
    test_certificate_analysis_with_heirs()
    print("\nAll tests passed successfully!")