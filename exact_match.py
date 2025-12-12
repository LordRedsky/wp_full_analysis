import io
from typing import Tuple, Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
BLUE_FILL = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
GREEN_FILL = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
GREY_FILL = PatternFill(start_color="FFBFBFBF", end_color="FFBFBFBF", fill_type="solid")


def _read_excel_both(file_bytes: bytes) -> Tuple[pd.DataFrame, Any]:
    df = pd.read_excel(io.BytesIO(file_bytes))
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    return df, wb


def _normalize_str(x: Any) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()


def _to_number(x: Any) -> float:
    if pd.isna(x):
        return float("nan")
    s = str(x)
    s = s.replace("Rp", "").replace(",", "").replace(".", "").replace(" ", "")
    try:
        return float(s)
    except Exception:
        try:
            return float(str(x))
        except Exception:
            return float("nan")

def _to_digit_str(x: Any) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    s = s.replace("Rp", "").replace(" ", "")
    digits = "".join(ch for ch in s if ch.isdigit())
    return digits

def _to_int(x: Any):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    try:
        if isinstance(x, (int, float)):
            return int(float(x))
        s = str(x).strip()
        s = s.replace("Rp", "").replace(" ", "").replace(",", "")
        if s.count(".") <= 1:
            try:
                return int(float(s))
            except Exception:
                pass
        digits = "".join(ch for ch in s if ch.isdigit())
        return int(digits) if digits else None
    except Exception:
        return None

def _parse_currency_to_int(x: Any):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    try:
        s = str(x).strip()
        s = s.replace("Rp", "").replace(" ", "").replace(".", "").replace(",", "")
        return int(s) if s.isdigit() else int(float(s))
    except Exception:
        return None

def _set_number(ws, row: int, col: int, value: Any):
    ws.cell(row=row, column=col, value=value)
    ws.cell(row=row, column=col).number_format = "0"

def _set_padded_number(ws, row: int, col: int, value: Any, width: int):
    n = _to_int(value)
    if n is None:
        ws.cell(row=row, column=col, value=str(value))
    else:
        ws.cell(row=row, column=col, value=n)
        ws.cell(row=row, column=col).number_format = "0" * width

def _set_text(ws, row: int, col: int, value: Any):
    ws.cell(row=row, column=col, value=str(value) if value is not None else None)
    ws.cell(row=row, column=col).number_format = "@"

def _set_text_padded(ws, row: int, col: int, value: Any, width: int):
    s = _normalize_str(value)
    if s == "":
        ws.cell(row=row, column=col, value="")
    else:
        if len(s) < width:
            s = s.zfill(width)
        ws.cell(row=row, column=col, value=s)
    ws.cell(row=row, column=col).number_format = "@"

def _zero_run_width(fmt: str) -> int:
    if not fmt:
        return 0
    max_run = 0
    run = 0
    for ch in fmt:
        if ch == "0":
            run += 1
            if run > max_run:
                max_run = run
        else:
            run = 0
    return max_run


def _find_header_col(ws, header_name: str) -> int:
    header_name_norm = header_name.lower().strip()
    for cell in ws[1]:
        if str(cell.value).lower().strip() == header_name_norm:
            return cell.column
    raise KeyError(f"Kolom '{header_name}' tidak ditemukan")


def _is_unfilled(cell) -> bool:
    return cell.fill is None or cell.fill.patternType is None


def _append_ket(ws, row: int, col: int, msg: str) -> None:
    existing = ws.cell(row=row, column=col).value
    if existing is None or str(existing).strip() == "":
        ws.cell(row=row, column=col, value=f"1. {msg}")
    else:
        lines = [l for l in str(existing).splitlines() if l.strip()]
        n = len(lines) + 1
        ws.cell(row=row, column=col, value=str(existing) + "\n" + f"{n}. {msg}")

def _normalize_name_for_match(name: str) -> str:
    s = name.lower().strip()
    s = " ".join(s.split())
    return s

def _contains_title(name: str) -> bool:
    s = name.lower()
    tokens = [t.replace(".", "") for t in s.replace(",", " ").split()]
    titles = {
        "dr", "drs", "ir", "h", "hj", "kh",
        "ssos", "sag", "spd", "sh", "se", "skom", "sip", "spsi", "skep", "st",
        "msi", "msos", "mkom", "mpd", "mh", "me", "mak", "mt", "msc", "ma", "phd",
        "sp", "spa", "spb", "spog"
    }
    for t in tokens:
        if t in titles:
            return True
    return False

def _contains_prefix(name: str) -> bool:
    s = name.lower().replace(",", " ")
    tokens = [t.replace(".", "") for t in s.split()]
    for i, t in enumerate(tokens):
        if t in {"laode", "waode", "muh", "m", "ld", "wd", "la", "wa"}:
            return True
        if t == "la" and i + 1 < len(tokens) and tokens[i + 1] == "ode":
            return True
        if t == "wa" and i + 1 < len(tokens) and tokens[i + 1] == "ode":
            return True
    return False

def _strip_titles_and_prefixes(name: str) -> str:
    s = name.lower().replace(",", " ").replace(".", " ")
    tokens = [t for t in s.split() if t]
    titles = {
        "dr", "drs", "ir", "h", "hj", "kh",
        "ssos", "sag", "spd", "sh", "se", "skom", "sip", "spsi", "skep", "st",
        "msi", "msos", "mkom", "mpd", "mh", "me", "mak", "mt", "msc", "ma", "phd",
        "sp", "spa", "spb", "spog"
    }
    prefixes = {"laode", "waode", "muh", "m", "ld", "wd", "la", "wa"}
    out = []
    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in titles or t in prefixes:
            i += 1
            continue
        if t in {"la", "wa"} and i + 1 < len(tokens) and tokens[i + 1] == "ode":
            i += 2
            continue
        out.append(t)
        i += 1
    return " ".join(out).strip()


def process_exact_match(
    wajib_pajak_file: bytes,
    sertifikat_file: bytes,
    kendali_file: bytes,
) -> Tuple[bytes, bytes, Dict[str, int]]:
    wp_df, wp_wb = _read_excel_both(wajib_pajak_file)
    cert_df, cert_wb = _read_excel_both(sertifikat_file)
    kendali_df, kendali_wb = _read_excel_both(kendali_file)
    kendali_df_orig = kendali_df.copy()

    wp_ws = wp_wb.active
    cert_ws = cert_wb.active
    kendali_ws = kendali_wb.active

    for col in ["NIB", "Luas"]:
        if col in cert_df.columns:
            if col == "NIB":
                cert_df[col] = cert_df[col].apply(_to_int)
            else:
                cert_df[col] = cert_df[col].apply(_to_int)

    for col in ["NIB", "LUASTERTUL", "RPBULAT"]:
        if col in kendali_df.columns:
            if col == "NIB":
                kendali_df[col] = kendali_df[col].apply(_to_int)
            elif col == "LUASTERTUL":
                kendali_df[col] = kendali_df[col].apply(_to_int)
            elif col == "RPBULAT":
                kendali_df[col] = kendali_df[col].apply(_parse_currency_to_int)

    # Build display mapping from original Kendali sheet
    kendali_display_by_nib = {}
    try:
        kendali_nib_col = _find_header_col(kendali_ws, "NIB")
        kendali_kls_col = _find_header_col(kendali_ws, "KELAS_BUMI")
        for r in range(2, kendali_ws.max_row + 1):
            raw_nib = kendali_ws.cell(row=r, column=kendali_nib_col).value
            nib_key = _to_int(raw_nib)
            if nib_key is None:
                continue
            raw_kls = kendali_ws.cell(row=r, column=kendali_kls_col).value
            nib_fmt = kendali_ws.cell(row=r, column=kendali_nib_col).number_format or ""
            kls_fmt = kendali_ws.cell(row=r, column=kendali_kls_col).number_format or ""
            w_n = _zero_run_width(nib_fmt)
            w_k = _zero_run_width(kls_fmt)
            nib_disp = str(raw_nib) if raw_nib is not None else ""
            kls_disp = str(raw_kls) if raw_kls is not None else ""
            if w_n > 0 and _to_int(raw_nib) is not None:
                nib_disp = str(_to_int(raw_nib)).zfill(w_n)
            if w_k > 0 and _to_int(raw_kls) is not None:
                kls_disp = str(_to_int(raw_kls)).zfill(w_k)
            kendali_display_by_nib[nib_key] = {"NIB": nib_disp, "KELAS_BUMI": kls_disp}
    except Exception:
        kendali_display_by_nib = {}

    wp_name_col = _find_header_col(wp_ws, "NAMA_WP")
    wp_jlh_col = _find_header_col(wp_ws, "JENIS_LHN")
    wp_kls_col = _find_header_col(wp_ws, "KELAS_BUMI")
    wp_nib1_col = _find_header_col(wp_ws, "NIB_1")
    wp_kdznt_col = _find_header_col(wp_ws, "KD_ZNT")
    wp_njop_col = _find_header_col(wp_ws, "NJOP_BUMI")
    wp_luas_col = _find_header_col(wp_ws, "LUAS_BUMI")

    cert_nama_col = _find_header_col(cert_ws, "Nama")
    cert_nib_col = _find_header_col(cert_ws, "NIB")
    cert_luas_col = _find_header_col(cert_ws, "Luas")

    try:
        cert_ket_col = _find_header_col(cert_ws, "KETERANGAN")
    except KeyError:
        cert_ket_col = cert_ws.max_column + 1
        cert_ws.cell(row=1, column=cert_ket_col, value="KETERANGAN")

    ahliwaris_count = 0
    for r in range(2, cert_ws.max_row + 1):
        nama_val_raw = cert_ws.cell(row=r, column=cert_nama_col).value
        nama_val = _normalize_str(nama_val_raw)
        if not nama_val:
            continue
        if len(nama_val) > 35:
            for c in [cert_nama_col, cert_nib_col, cert_luas_col]:
                cert_ws.cell(row=r, column=c).fill = YELLOW_FILL
            _append_ket(cert_ws, r, cert_ket_col, "Data ahli waris")
            ahliwaris_count += 1

    nama_to_cert_rows: Dict[str, list] = {}
    base_to_cert_rows: Dict[str, list] = {}
    for r in range(2, cert_ws.max_row + 1):
        nama_cell = cert_ws.cell(row=r, column=cert_nama_col)
        if not _is_unfilled(nama_cell):
            continue
        nama_val = _normalize_str(nama_cell.value)
        if not nama_val:
            continue
        nib_val = cert_ws.cell(row=r, column=cert_nib_col).value
        luas_val = cert_ws.cell(row=r, column=cert_luas_col).value
        item = {"row": r, "NIB": _to_int(nib_val), "Luas": _to_int(luas_val), "Nama": nama_val}
        key = _normalize_name_for_match(nama_val)
        base_key = _strip_titles_and_prefixes(nama_val)
        nama_to_cert_rows.setdefault(key, []).append(item)
        if base_key:
            base_to_cert_rows.setdefault(base_key, []).append(item)

    kendali_by_nib = {}
    if "NIB" in kendali_df.columns:
        for i, row in kendali_df.iterrows():
            nib = _to_int(row.get("NIB"))
            if nib is None:
                continue
            disp = kendali_display_by_nib.get(nib, {})
            kendali_by_nib[nib] = {
                "RPBULAT": _parse_currency_to_int(row.get("RPBULAT")),
                "JENIS_ZONA": _normalize_str(row.get("JENIS_ZONA")),
                "KELAS_BUMI": _normalize_str(row.get("KELAS_BUMI")),
                "KD_ZNT": _normalize_str(row.get("KD_ZNT")),
                "LUASTERTUL": _to_int(row.get("LUASTERTUL")),
                "NIB_DISPLAY": disp.get("NIB", _normalize_str(kendali_df_orig.loc[i, "NIB"]) if "NIB" in kendali_df_orig.columns else ""),
                "KELAS_BUMI_DISPLAY": disp.get("KELAS_BUMI", _normalize_str(kendali_df_orig.loc[i, "KELAS_BUMI"]) if "KELAS_BUMI" in kendali_df_orig.columns else ""),
            }

    try:
        wp_ket_col = _find_header_col(wp_ws, "KETERANGAN")
    except KeyError:
        wp_ket_col = wp_ws.max_column + 1
        wp_ws.cell(row=1, column=wp_ket_col, value="KETERANGAN")

    processed_count = 0
    matched_count = 0
    nib_not_found_count = 0
    name_not_found_count = 0
    for r in range(2, wp_ws.max_row + 1):
        nama_wp = _normalize_str(wp_ws.cell(row=r, column=wp_name_col).value)
        if not nama_wp:
            continue
        processed_count += 1
        candidates = nama_to_cert_rows.get(_normalize_name_for_match(nama_wp))
        if not candidates:
            base_wp = _strip_titles_and_prefixes(nama_wp)
            wp_decor = _contains_title(nama_wp) or _contains_prefix(nama_wp)
            base_candidates = base_to_cert_rows.get(base_wp) if base_wp else None
            chosen = None
            matched_via_base = False
            if base_candidates:
                filtered = []
                for it in base_candidates:
                    cert_decor = _contains_title(it["Nama"]) or _contains_prefix(it["Nama"]) 
                    if wp_decor != cert_decor and _is_unfilled(cert_ws.cell(row=it["row"], column=cert_nama_col)):
                        filtered.append(it)
                for it in filtered:
                    nib = it["NIB"]
                    if nib is None:
                        continue
                    kendali = kendali_by_nib.get(nib)
                    if kendali:
                        chosen = (it, kendali)
                        matched_via_base = True
                        break
            if not chosen:
                if _contains_title(nama_wp) or _contains_prefix(nama_wp):
                    wp_ws.cell(row=r, column=wp_name_col).fill = GREY_FILL
                    _append_ket(wp_ws, r, wp_ket_col, "Nama Wajib Pajak tidak ditemukan di sertifikat")
                else:
                    wp_ws.cell(row=r, column=wp_name_col).fill = RED_FILL
                    _append_ket(wp_ws, r, wp_ket_col, "Nama Wajib Pajak tidak ditemukan di sertifikat")
                name_not_found_count += 1
                continue
        else:
            chosen = None
            matched_via_base = False
        if candidates and chosen is None:
            for item in candidates:
                row_chk = item["row"]
                if not _is_unfilled(cert_ws.cell(row=row_chk, column=cert_nama_col)):
                    continue
                nib = item["NIB"]
                if nib is None:
                    continue
                kendali = kendali_by_nib.get(nib)
                if kendali:
                    chosen = (item, kendali)
                    break
        if candidates and not chosen:
            white_candidates = [i for i in candidates if _is_unfilled(cert_ws.cell(row=i["row"], column=cert_nama_col))]
            if not white_candidates:
                if _contains_title(nama_wp) or _contains_prefix(nama_wp):
                    wp_ws.cell(row=r, column=wp_name_col).fill = GREY_FILL
                    _append_ket(wp_ws, r, wp_ket_col, "Nama Wajib Pajak tidak ditemukan di sertifikat")
                else:
                    wp_ws.cell(row=r, column=wp_name_col).fill = RED_FILL
                    _append_ket(wp_ws, r, wp_ket_col, "Nama Wajib Pajak tidak ditemukan di sertifikat")
                name_not_found_count += 1
                continue
            item = white_candidates[0]
            cert_row = item["row"]
            wp_ws.cell(row=r, column=wp_name_col).fill = RED_FILL
            for c in [cert_nama_col, cert_nib_col, cert_luas_col]:
                cert_ws.cell(row=cert_row, column=c).fill = RED_FILL
            _append_ket(cert_ws, cert_row, cert_ket_col, "NIB tidak ditemukan pada form kendali")
            _append_ket(wp_ws, r, wp_ket_col, "NIB tidak ditemukan pada form kendali")
            _set_padded_number(wp_ws, r, wp_nib1_col, item["NIB"], 5)
            nib_not_found_count += 1
            continue

        item, kendali = chosen
        cert_row = item["row"]
        cert_luas = item["Luas"]
        nib = item["NIB"]

        wp_ws.cell(row=r, column=wp_jlh_col, value=kendali["JENIS_ZONA"]) 
        _set_text_padded(wp_ws, r, wp_kls_col, kendali.get("KELAS_BUMI_DISPLAY", kendali["KELAS_BUMI"]), 3) 
        _set_text_padded(wp_ws, r, wp_nib1_col, kendali.get("NIB_DISPLAY", nib), 5)
        wp_ws.cell(row=r, column=wp_kdznt_col, value=kendali["KD_ZNT"]) 
        njop = kendali["RPBULAT"]
        if njop is not None:
            _set_number(wp_ws, r, wp_njop_col, njop)
        else:
            wp_ws.cell(row=r, column=wp_njop_col, value=njop)

        luas_tertul = kendali["LUASTERTUL"]
        old_wp_luas = wp_ws.cell(row=r, column=wp_luas_col).value
        old_norm = _to_int(old_wp_luas)
        tertul_norm = _to_int(luas_tertul)
        if tertul_norm is None:
            pass
        else:
            if old_norm != tertul_norm:
                _set_number(wp_ws, r, wp_luas_col, tertul_norm)
                wp_ws.cell(row=r, column=wp_luas_col).fill = GREEN_FILL
                _append_ket(wp_ws, r, wp_ket_col, f"Telah dilakukan penyesuaian luas bumi sesuai data sertipikat. {old_wp_luas} -> {luas_tertul}")

        for c in [cert_nama_col, cert_nib_col, cert_luas_col]:
            cert_ws.cell(row=cert_row, column=c).fill = BLUE_FILL
        _append_ket(cert_ws, cert_row, cert_ket_col, "Data cocok")
        matched_count += 1

        if matched_via_base:
            old_name = wp_ws.cell(row=r, column=wp_name_col).value
            new_name = item["Nama"]
            wp_ws.cell(row=r, column=wp_name_col, value=new_name)
            wp_ws.cell(row=r, column=wp_name_col).fill = GREEN_FILL
            _append_ket(wp_ws, r, wp_ket_col, f"Telah dilakukan penyesuaian nama sesuai sertifikat. {old_name} -> {new_name}")

    wp_out = io.BytesIO()
    cert_out = io.BytesIO()
    wp_wb.save(wp_out)
    cert_wb.save(cert_out)
    stats = {
        "processed": processed_count,
        "matched": matched_count,
        "nib_not_found": nib_not_found_count,
        "name_not_found": name_not_found_count,
        "ahli_waris": ahliwaris_count,
    }
    return wp_out.getvalue(), cert_out.getvalue(), stats
