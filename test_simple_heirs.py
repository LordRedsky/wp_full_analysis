#!/usr/bin/env python3
"""
Simple test to verify heirs detection with properly white-filled cells
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from analysis_sert import process_certificate_analysis, _is_unfilled


def create_simple_test():
    """Create a simple test file with clearly white-filled cells"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add headers
    ws.cell(row=1, column=1, value="Nama")
    ws.cell(row=1, column=2, value="NIB")
    ws.cell(row=1, column=3, value="Luas")

    # Row 2: Multiple names, both NIB and NAMA should be white-filled
    ws.cell(row=2, column=1, value="HAFIDIN, A.MA.PD.I, Drs. Amin")
    ws.cell(row=2, column=2, value="12345")
    ws.cell(row=2, column=3, value=100)
    
    # Verify the cells are white-filled (no fill applied)
    print(f"Row 2 - NIB cell fill type: {ws.cell(row=2, column=2).fill.fill_type}")
    print(f"Row 2 - NAMA cell fill type: {ws.cell(row=2, column=1).fill.fill_type}")
    print(f"Row 2 - NIB is unfilled: {_is_unfilled(ws.cell(row=2, column=2))}")
    print(f"Row 2 - NAMA is unfilled: {_is_unfilled(ws.cell(row=2, column=1))}")

    # Row 3: Single name, both NIB and NAMA should be white-filled
    ws.cell(row=3, column=1, value="John Doe")
    ws.cell(row=3, column=2, value="67890")
    ws.cell(row=3, column=3, value=200)
    
    print(f"Row 3 - NIB cell fill type: {ws.cell(row=3, column=2).fill.fill_type}")
    print(f"Row 3 - NAMA cell fill type: {ws.cell(row=3, column=1).fill.fill_type}")
    print(f"Row 3 - NIB is unfilled: {_is_unfilled(ws.cell(row=3, column=2))}")
    print(f"Row 3 - NAMA is unfilled: {_is_unfilled(ws.cell(row=3, column=1))}")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def test_simple_heirs_detection():
    """Test heirs detection with simple case"""
    print("Creating simple test file...")
    test_file = create_simple_test()

    print("\nRunning certificate analysis...")
    result = process_certificate_analysis(test_file)

    print("Certificate analysis completed!")
    print(f"Data diproses: {result['data_diproses']}")
    print(f"NIB duplikat: {result['nib_duplikat']}")
    print(f"Jumlah Ahli Waris: {result['jumlah_ahli_waris']}")

    # Load the result to check colors
    from openpyxl import load_workbook
    result_wb = load_workbook(io.BytesIO(result['workbook_bytes']))
    result_ws = result_wb.active

    print("\nAnalyzing results:")
    for r in range(2, 4):  # Check rows 2 and 3
        nama = result_ws.cell(row=r, column=1).value
        nib = result_ws.cell(row=r, column=2).value
        
        # Check colors
        nama_fill = result_ws.cell(row=r, column=1).fill
        nib_fill = result_ws.cell(row=r, column=2).fill
        luas_fill = result_ws.cell(row=r, column=3).fill

        is_red = (hasattr(nama_fill, 'start_color') and 
                 nama_fill.start_color.rgb == "FFFF0000") if hasattr(nama_fill, 'start_color') else False
        is_yellow = (hasattr(nama_fill, 'start_color') and 
                    nama_fill.start_color.rgb == "FFFFFF00") if hasattr(nama_fill, 'start_color') else False

        print(f"Row {r}: Nama='{nama}', NIB='{nib}'")
        print(f"  Is Red: {is_red}, Is Yellow: {is_yellow}")
        
        # Check if it's the KETERANGAN column for any notes
        # Find KETERANGAN column
        ket_col = None
        for col in range(1, result_ws.max_column + 1):
            if result_ws.cell(row=1, column=col).value == "KETERANGAN":
                ket_col = col
                break
        
        if ket_col:
            ket_value = result_ws.cell(row=r, column=ket_col).value
            print(f"  KETERANGAN: {ket_value}")

    print(f"\nTotal heirs detected: {result['jumlah_ahli_waris']}")


if __name__ == "__main__":
    test_simple_heirs_detection()