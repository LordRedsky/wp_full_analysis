#!/usr/bin/env python3
"""
Edge case test for heirs detection to ensure we don't miss any cases
"""

import io
from openpyxl import Workbook
from analysis_sert import process_certificate_analysis


def test_edge_cases():
    """Test edge cases for heirs detection"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Add headers
    ws.cell(row=1, column=1, value="Nama")
    ws.cell(row=1, column=2, value="NIB")
    ws.cell(row=1, column=3, value="Luas")

    # Row 2: Clear multiple names - should be yellow
    ws.cell(row=2, column=1, value="A, B, C")
    ws.cell(row=2, column=2, value="11111")
    ws.cell(row=2, column=3, value=100)

    # Row 3: Multiple names with various titles - should be yellow
    ws.cell(row=3, column=1, value="Siti, H. Ahmad, Drs. Budi, Ir. Candra")
    ws.cell(row=3, column=2, value="22222")
    ws.cell(row=3, column=3, value=200)

    # Row 4: Single name with comma in title - should NOT be yellow
    ws.cell(row=4, column=1, value="Dr., Ir. Joko")
    ws.cell(row=4, column=2, value="33333")
    ws.cell(row=4, column=3, value=300)

    # Row 5: Multiple names with complex titles - should be yellow
    ws.cell(row=5, column=1, value="H. Joko, A.Md.Kom, Siti, H. Ahmad")
    ws.cell(row=5, column=2, value="44444")
    ws.cell(row=5, column=3, value=400)

    # Row 6: Name with initials that might be confused - should NOT be yellow
    ws.cell(row=6, column=1, value="H. A. Budi")
    ws.cell(row=6, column=2, value="55555")
    ws.cell(row=6, column=3, value=500)

    output = io.BytesIO()
    wb.save(output)
    test_file = output.getvalue()

    print("Running edge case test...")
    result = process_certificate_analysis(test_file)

    print(f"Jumlah Ahli Waris: {result['jumlah_ahli_waris']}")
    
    # Load the result to check colors
    from openpyxl import load_workbook
    result_wb = load_workbook(io.BytesIO(result['workbook_bytes']))
    result_ws = result_wb.active

    print("\nAnalyzing edge case results:")
    for r in range(2, 7):  # Check rows 2-6
        nama = result_ws.cell(row=r, column=1).value
        nib = result_ws.cell(row=r, column=2).value
        
        # Check colors
        nama_fill = result_ws.cell(row=r, column=1).fill
        is_red = (hasattr(nama_fill, 'start_color') and 
                 nama_fill.start_color.rgb == "FFFF0000") if hasattr(nama_fill, 'start_color') else False
        is_yellow = (hasattr(nama_fill, 'start_color') and 
                    nama_fill.start_color.rgb == "FFFFFF00") if hasattr(nama_fill, 'start_color') else False

        # Check KETERANGAN
        ket_col = None
        for col in range(1, result_ws.max_column + 1):
            if result_ws.cell(row=1, column=col).value == "KETERANGAN":
                ket_col = col
                break
        
        ket_value = result_ws.cell(row=r, column=ket_col).value if ket_col else "No KET col"
        
        print(f"Row {r}: '{nama}' -> NIB={nib}")
        print(f"  Is Red: {is_red}, Is Yellow: {is_yellow}, KET: {ket_value}")

    # Expected results:
    # - Row 2: A, B, C -> 3 names -> yellow ✓
    # - Row 3: Siti, Ahmad, Budi, Candra -> 4 names -> yellow ✓
    # - Row 4: Dr., Ir. Joko -> 1 name after parsing -> NOT yellow ✓
    # - Row 5: Joko, Siti, Ahmad -> 3 names -> yellow ✓
    # - Row 6: H. A. Budi -> likely just "Budi" after parsing -> NOT yellow ✓
    
    expected_heirs = 3  # Rows 2, 3, and 5
    print(f"\nExpected heirs: {expected_heirs}, Found: {result['jumlah_ahli_waris']}")
    
    if result['jumlah_ahli_waris'] == expected_heirs:
        print("✓ Edge case test passed!")
    else:
        print("✗ Edge case test failed!")


if __name__ == "__main__":
    test_edge_cases()